# -*- coding: utf-8 -*-
##Pyslvs - Open Source Planar Linkage Mechanism Simulation and Dimensional Synthesis System.
##Copyright (C) 2016-2017 Yuan Chang
##E-mail: pyslvs@gmail.com
##
##This program is free software; you can redistribute it and/or modify
##it under the terms of the GNU Affero General Public License as published by
##the Free Software Foundation; either version 3 of the License, or
##(at your option) any later version.
##
##This program is distributed in the hope that it will be useful,
##but WITHOUT ANY WARRANTY; without even the implied warranty of
##MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
##GNU Affero General Public License for more details.
##
##You should have received a copy of the GNU Affero General Public License
##along with this program; if not, write to the Free Software
##Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA 02111-1307, USA.

from ...QtModules import *
from .Ui_Algorithm import Ui_Form as PathSolving_Form
from ...graphics.ChartGraphics import ChartDialog
from ...graphics.Algorithm_preview import PreviewDialog
from ...libs.pyslvs_algorithm.TS import solver, Direction
from .Algorithm_options import Algorithm_options_show
from .Algorithm_path_adjust import Algorithm_path_adjust_show
from .Algorithm_progress import Algorithm_progress_show
from .Algorithm_series import Algorithm_series_show
import csv
import openpyxl
from re import split as charSplit

class Algorithm_show(QWidget, PathSolving_Form):
    fixPointRange = pyqtSignal(tuple, float, tuple, float)
    pathChanged = pyqtSignal(tuple)
    mergeResult = pyqtSignal(int, tuple, tuple)
    GeneticPrams = {'nPop':500, 'pCross':0.95, 'pMute':0.05, 'pWin':0.95, 'bDelta':5.}
    FireflyPrams = {'n':80, 'alpha':0.01, 'betaMin':0.2, 'gamma':1., 'beta0':1.}
    DifferentialPrams = {'strategy':1, 'NP':400, 'F':0.6, 'CR':0.9}
    defaultSettings = {
        'maxGen':1500, 'report':1, 'IMin':5., 'LMin':5., 'FMin':5., 'AMin':0.,
        'IMax':100., 'LMax':100., 'FMax':100., 'AMax':360., 'algorithmPrams':DifferentialPrams
    }
    mechanismParams_4Bar = { #No 'targetPath'
        'Driving':'A',
        'Follower':'D',
        'Link':'L0,L1,L2,L3,L4',
        'Target':'E',
        'ExpressionName':'PLAP,PLLP,PLLP',
        'Expression':'A,L0,a0,D,B,B,L1,L2,D,C,B,L3,L4,C,E',
        'constraint':[{'driver':'L0', 'follower':'L2', 'connect':'L1'}]
    }
    mechanismParams_4Bar['VARS'] = len(set(mechanismParams_4Bar['Expression'].split(',')))-2
    mechanismParams_8Bar = { #No 'targetPath'
        'Driving':'A',
        'Follower':'B',
        'Link':'L0,L1,L2,L3,L4,L5,L6,L7,L8,L9,L10',
        'Target':'H',
        'ExpressionName':'PLAP,PLLP,PLLP,PLLP,PLLP,PLLP',
        'Expression':'A,L0,a0,B,C,B,L2,L1,C,D,B,L4,L3,D,E,C,L5,L6,B,F,F,L8,L7,E,G,F,L9,L10,G,H',
        'constraint':[{'driver':'L0', 'follower':'L2', 'connect':'L1'}]
    }
    mechanismParams_8Bar['VARS'] = len(set(mechanismParams_8Bar['Expression'].split(',')))-2
    
    def __init__(self, parent):
        super(Algorithm_show, self).__init__(parent)
        self.setupUi(self)
        self.path = parent.FileWidget.Designs.path
        self.mechanism_data = parent.FileWidget.Designs.result
        self.mechanism_data_add = parent.FileWidget.Designs.addResult
        self.mechanism_data_del = parent.FileWidget.Designs.delResult
        self.env = lambda: parent.env
        self.unsaveFunc = parent.workbookNoSave
        self.Settings = self.defaultSettings
        self.algorithmPrams_default()
        self.Point_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.Point_list.customContextMenuRequested.connect(self.on_Point_list_context_menu)
        self.popMenu_list = QMenu(self)
        self.action_paste_from_clipboard = QAction("&Paste from clipboard", self)
        self.popMenu_list.addAction(self.action_paste_from_clipboard)
        self.Ar.valueChanged.connect(self.updateRange)
        self.Ax.valueChanged.connect(self.updateRange)
        self.Ay.valueChanged.connect(self.updateRange)
        self.Dr.valueChanged.connect(self.updateRange)
        self.Dx.valueChanged.connect(self.updateRange)
        self.Dy.valueChanged.connect(self.updateRange)
        self.type0.clicked.connect(self.algorithmPrams_default)
        self.type1.clicked.connect(self.algorithmPrams_default)
        self.type2.clicked.connect(self.algorithmPrams_default)
        self.Result_list.clicked.connect(self.hasResult)
        self.isGenerate()
        self.hasResult()
    
    def loadResults(self):
        for e in self.mechanism_data:
            self.addResult(e)
    
    @pyqtSlot()
    def on_clearAll_clicked(self):
        self.Point_list.setCurrentRow(0)
        for i in reversed(range(self.Point_list.count()+1)):
            self.on_remove_clicked()
        self.isGenerate()
    
    @pyqtSlot()
    def on_series_clicked(self):
        dlg = Algorithm_series_show(self)
        dlg.show()
        if dlg.exec_():
            for e in dlg.path:
                self.on_add_clicked(e[0], e[1])
    
    def on_Point_list_context_menu(self, point):
        action = self.popMenu_list.exec_(self.Point_list.mapToGlobal(point))
        if action==self.action_paste_from_clipboard:
            data = QApplication.clipboard().text()
            data = charSplit(",|\n", data)
            self.readPathFromCSV(data)
    
    @pyqtSlot()
    def on_importCSV_clicked(self):
        fileName, _ = QFileDialog.getOpenFileName(self, "Open file...", self.env(), "Text File (*.txt);;CSV File (*.csv)")
        if fileName:
            data = []
            with open(fileName, newline=str()) as stream:
                reader = csv.reader(stream, delimiter=' ', quotechar='|')
                for row in reader:
                    data += ' '.join(row).split(',')
            self.readPathFromCSV(data)
    
    def readPathFromCSV(self, data):
        try:
            data = [(round(float(data[i]), 4), round(float(data[i+1]), 4)) for i in range(0, len(data), 2)]
            for e in data:
                self.on_add_clicked(e[0], e[1])
        except:
            dlgbox = QMessageBox(QMessageBox.Warning, "File error", "Wrong format.\nIt should be look like this:"+
                "\n0.0,0.0[\\n]"*3, (QMessageBox.Ok), self)
            dlgbox.exec_()
    
    @pyqtSlot()
    def on_importXLSX_clicked(self):
        fileName, _ = QFileDialog.getOpenFileName(self, "Open file...", self.env(), "Microsoft Office Excel (*.xlsx *.xlsm *.xltx *.xltm)")
        if fileName:
            wb = openpyxl.load_workbook(fileName)
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            data = []
            #Keep finding until there is no value.
            i = 1
            while True:
                x = ws.cell(row=i, column=1).value
                y = ws.cell(row=i, column=2).value
                if x==None or y==None:
                    break
                try:
                    data.append((round(float(x), 4), round(float(y), 4)))
                except:
                    dlgbox = QMessageBox(QMessageBox.Warning, "File error", "Wrong format.\nThe datasheet seems to including non-digital cell.", (QMessageBox.Ok), self)
                    dlgbox.exec_()
                    break
                i += 1
            for e in data:
                self.on_add_clicked(e[0], e[1])
    
    @pyqtSlot()
    def on_pathAdjust_clicked(self):
        dlg = Algorithm_path_adjust_show(self.path)
        dlg.show()
        if dlg.exec_():
            self.on_clearAll_clicked()
            for e in dlg.get_path():
                self.on_add_clicked(e[0], e[1])
    
    @pyqtSlot()
    def on_moveUp_clicked(self):
        row = self.Point_list.currentRow()
        if row > 0 and self.Point_list.count() > 1:
            self.path.insert(row-1, (self.path[row][0], self.path[row][1]))
            del self.path[row+1]
            self.pathChanged.emit(tuple(self.path))
            x = self.Point_list.currentItem().text()[1:-1].split(", ")[0]
            y = self.Point_list.currentItem().text()[1:-1].split(", ")[1]
            self.Point_list.insertItem(row-1, "({}, {})".format(x, y))
            self.Point_list.takeItem(row+1)
            self.Point_list.setCurrentRow(row-1)
    
    @pyqtSlot()
    def on_moveDown_clicked(self):
        row = self.Point_list.currentRow()
        if row < self.Point_list.count()-1 and self.Point_list.count() > 1:
            self.path.insert(row+2, (self.path[row][0], self.path[row][1]))
            del self.path[row]
            self.pathChanged.emit(tuple(self.path))
            c = self.Point_list.currentItem().text()[1:-1].split(", ")
            self.Point_list.insertItem(row+2, "({}, {})".format(c[0], c[1]))
            self.Point_list.takeItem(row)
            self.Point_list.setCurrentRow(row+1)
    
    def addPath(self, x, y):
        self.Point_list.addItem('({}, {})'.format(x, y))
        self.isGenerate()
    
    @pyqtSlot()
    def on_add_clicked(self, x=None, y=None):
        if x is None:
            x = self.X_coordinate.value()
            y = self.Y_coordinate.value()
        self.path.append((x, y))
        self.pathChanged.emit(tuple(self.path))
        self.Point_list.addItem("({}, {})".format(x, y))
        self.isGenerate()
    
    @pyqtSlot()
    def on_remove_clicked(self):
        if self.Point_list.currentRow()>-1:
            del self.path[self.Point_list.currentRow()]
            self.pathChanged.emit(tuple(self.path))
            self.Point_list.takeItem(self.Point_list.currentRow())
            self.isGenerate()
    
    @pyqtSlot()
    def on_close_path_clicked(self):
        if self.Point_list.count() > 1 and self.path[0]!=self.path[-1]:
            self.on_add_clicked(*self.path[0])
    
    def isGenerate(self):
        self.pointNum.setText(
            "<html><head/><body><p><span style=\"font-size:12pt; color:#00aa00;\">{}</span></p></body></html>".format(self.Point_list.count())
        )
        n = self.Point_list.count()>1
        self.pathAdjust.setEnabled(n)
        self.GenerateLocal.setEnabled(n)
        self.GenerateZMQ.setEnabled(n)
    
    @pyqtSlot()
    def on_GenerateLocal_clicked(self):
        self.startAlgorithm()
    
    @pyqtSlot()
    def on_GenerateZMQ_clicked(self):
        self.startAlgorithm(hasPort=True)
    
    def startAlgorithm(self, hasPort=False):
        type_num, mechanismParams, setting = self.getGenerate()
        setting.update(self.Settings['algorithmPrams'])
        dlg = Algorithm_progress_show(type_num, mechanismParams, setting, self.portText.text() if hasPort else None, self)
        dlg.show()
        if dlg.exec_():
            self.mechanism_data_add(dlg.mechanisms)
            for m in dlg.mechanisms:
                self.addResult(m)
            self.setTime(dlg.time_spand)
            self.unsaveFunc()
            dlgbox = QMessageBox(QMessageBox.Information, "Dimensional Synthesis", "Your tasks is all completed.", (QMessageBox.Ok), self.parent())
            if dlgbox.exec_():
                print("Finished.")
    
    def getGenerate(self):
        type_num = 0 if self.type0.isChecked() else 1 if self.type1.isChecked() else 2
        mechanismParams = self.mechanismParams_4Bar if self.FourBar.isChecked() else self.mechanismParams_8Bar
        link_q = mechanismParams['VARS']-7
        upper = [self.Ax.value()+self.Ar.value()/2, self.Ay.value()+self.Ar.value()/2, self.Dx.value()+self.Dr.value()/2, self.Dy.value()+self.Dr.value()/2,
            self.Settings['IMax'], self.Settings['LMax'], self.Settings['FMax']]+[self.Settings['LMax']]*link_q
        lower = [self.Ax.value()-self.Ar.value()/2, self.Ay.value()-self.Ar.value()/2, self.Dx.value()-self.Dr.value()/2, self.Dy.value()-self.Dr.value()/2,
            self.Settings['IMin'], self.Settings['LMin'], self.Settings['FMin']]+[self.Settings['LMin']]*link_q
        mechanismParams['targetPath'] = tuple(self.path)
        p = len(self.path)
        generateData = {
            'nParm':p+mechanismParams['VARS'],
            'upper':upper+[self.Settings['AMax']]*p,
            'lower':lower+[self.Settings['AMin']]*p,
            'maxGen':self.Settings['maxGen'],
            'report':int(self.Settings['maxGen']*self.Settings['report']/100)
        }
        return type_num, mechanismParams, generateData
    
    def setTime(self, time_spand):
        sec = round(time_spand%60, 2)
        mins = int(time_spand/60)
        self.timeShow.setText("<html><head/><body><p><span style=\"font-size:12pt\">{}[min] {}[s]</span></p></body></html>".format(mins, sec))
    
    #Add result items, except add to the list.
    def addResult(self, result):
        item = QListWidgetItem("{} ({} gen)".format(result['Algorithm'], result['settings']['maxGen']))
        interrupt = result['interrupted']
        if interrupt=='False':
            item.setIcon(QIcon(QPixmap(":/icons/task-completed.png")))
        elif interrupt=='N/A':
            item.setIcon(QIcon(QPixmap(":/icons/question-mark.png")))
        else:
            item.setIcon(QIcon(QPixmap(":/icons/interrupted.png")))
        keys = sorted(list(result.keys()))
        info = (["{}: {}".format(k, result[k]) for k in keys if 'x' in k or 'y' in k or 'L' in k]+
            ["\nClick to apply setting."]+["Double click to see dynamic preview."])
        item.setToolTip('\n'.join(["[{}] ({}{} gen)".format(
            result['Algorithm'],
            '' if interrupt=='False' else interrupt+'-',
            result['settings']['maxGen'])]+["※ Completeness is not clear." if interrupt=='N/A' else '']+info
        ))
        self.Result_list.addItem(item)
    
    @pyqtSlot()
    def on_deleteButton_clicked(self):
        row = self.Result_list.currentRow()
        self.mechanism_data_del(row)
        self.Result_list.takeItem(row)
        self.unsaveFunc()
        self.hasResult()
    
    def hasResult(self, p0=None):
        for button in [self.mergeButton, self.deleteButton]:
            button.setEnabled(self.Result_list.currentRow()>-1)
    
    @pyqtSlot(QModelIndex)
    def on_Result_list_doubleClicked(self, index):
        row = self.Result_list.currentRow()
        if row>-1:
            mechanism = self.mechanism_data[row]
            _, Paths = self.legal_crank()
            dlg = PreviewDialog(mechanism, Paths, self)
            dlg.show()
            dlg.exec_()
    
    @pyqtSlot()
    def on_mergeButton_clicked(self):
        row = self.Result_list.currentRow()
        if row>-1:
            reply = QMessageBox.question(self, "Message", "Merge this result to your canvas?",
                (QMessageBox.Apply | QMessageBox.Cancel), QMessageBox.Apply)
            if reply==QMessageBox.Apply:
                self.mergeResult.emit(row, *self.legal_crank())
    
    def legal_crank(self):
        row = self.Result_list.currentRow()
        Result = self.mechanism_data[row]
        path = Result['targetPath']
        pointAvg = sum([e[1] for e in path])/len(path)
        other = (Result['Ay']+Result['Dy'])/2>pointAvg and Result['Ax']<Result['Dx']
        answer = [False]
        expression = (self.mechanismParams_8Bar if Result['type']=='8Bar' else self.mechanismParams_4Bar)['Expression'].split(',')
        '''
        expression_tag
        four_bar = (
            ('A', 'L0', 'a0', 'D', 'B'),
            ('B', 'L1', 'L2', 'D', 'C'),
            ('B', 'L3', 'L4', 'C', 'E')
        )
        eight_bar = (
            ('A', 'L0', 'a0', 'B', 'C'),
            ('B', 'L1', 'L2', 'C', 'D'),
            ('B', 'L3', 'L4', 'D', 'E'),
            ('C', 'L5', 'L6', 'B', 'F'),
            ('F', 'L7', 'L8', 'E', 'G'),
            ('F', 'L9', 'L10', 'G', 'H')
        )
        '''
        expression_tag = tuple(tuple(expression[i+j] for j in range(5)) for i in range(0, len(expression), 5))
        expression_result = [exp[-1] for exp in expression_tag]
        exp_symbol = (expression_tag[0][0], expression_tag[0][3])+tuple(exp[-1] for exp in expression_tag)
        '''
        ('A', 'D', 'B', 'C', 'E')
        '''
        Paths = tuple([] for tag in exp_symbol)
        for a in range(360+1):
            Directions = [Direction(p1=(Result['Ax'], Result['Ay']), p2=(Result['Ax']+10, Result['Ay']), len1=Result['L0'], angle=a, other=other)]
            for exp in expression_tag[1:]:
                p1 = (Result['Ax'], Result['Ay']) if exp[0]=='A' else expression_result.index(exp[0]) if exp[0] in expression_result else (Result['Dx'], Result['Dy'])
                p2 = (Result['Ax'], Result['Ay']) if exp[3]=='A' else expression_result.index(exp[3]) if exp[3] in expression_result else (Result['Dx'], Result['Dy'])
                Directions.append(Direction(p1=p1, p2=p2, len1=Result[exp[1]], len2=Result[exp[2]], other=other))
            s = solver(Directions)
            s_answer = s.answer()
            if False not in s_answer:
                answer = [(Result['Ax'], Result['Ay']), (Result['Dx'], Result['Dy'])]+s_answer
            for i, a in enumerate(s_answer):
                Paths[exp_symbol.index(expression_result[i])].append(a)
        return tuple(answer), tuple(tuple(path) if (path==False or len(set(path))>1 or (False in path)) else () for path in Paths)
    
    @pyqtSlot()
    def on_Result_chart_clicked(self):
        dlg = ChartDialog("Convergence Value", self.mechanism_data, self)
        dlg.show()
    
    @pyqtSlot()
    def on_Result_load_settings_clicked(self):
        self.hasResult()
        row = self.Result_list.currentRow()
        if row>-1 and row!=len(self.mechanism_data):
            args = self.mechanism_data[row]
            if args['Algorithm']=='RGA':
                self.type0.setChecked(True)
            elif args['Algorithm']=='Firefly':
                self.type1.setChecked(True)
            elif args['Algorithm']=='DE':
                self.type2.setChecked(True)
            if args['type']=='8Bar':
                self.EightBar.setChecked(True)
            else:
                self.FourBar.setChecked(True)
            self.setTime(args['time'])
            settings = args['settings']
            self.Ax.setValue((settings['upper'][0]+settings['lower'][0])/2)
            self.Ay.setValue((settings['upper'][1]+settings['lower'][1])/2)
            self.Ar.setValue(abs(settings['upper'][0]-self.Ax.value())*2)
            self.Dx.setValue((settings['upper'][2]+settings['lower'][2])/2)
            self.Dy.setValue((settings['upper'][3]+settings['lower'][3])/2)
            self.Dr.setValue(abs(settings['upper'][2]-self.Dx.value())*2)
            self.Settings = {
                'maxGen':settings['maxGen'],
                'report':0 if settings['report']==0 else settings['maxGen']/settings['report']/100,
                'IMax':settings['upper'][4], 'IMin':settings['lower'][4],
                'LMax':settings['upper'][5], 'LMin':settings['lower'][5],
                'FMax':settings['upper'][6], 'FMin':settings['lower'][6],
                'AMax':settings['upper'][-1], 'AMin':settings['lower'][-1]
            }
            algorithmPrams = settings.copy()
            del algorithmPrams['nParm']
            del algorithmPrams['upper']
            del algorithmPrams['lower']
            del algorithmPrams['report']
            self.Settings['algorithmPrams'] = algorithmPrams
            self.on_clearAll_clicked()
            for e in args['targetPath']:
                self.on_add_clicked(e[0], e[1])
    
    def algorithmPrams_default(self):
        type_num = 0 if self.type0.isChecked() else 1 if self.type1.isChecked() else 2
        if type_num==0:
            self.Settings['algorithmPrams'] = self.GeneticPrams
        elif type_num==1:
            self.Settings['algorithmPrams'] = self.FireflyPrams
        elif type_num==2:
            self.Settings['algorithmPrams'] = self.DifferentialPrams
    
    @pyqtSlot()
    def on_advanceButton_clicked(self):
        type_num = "Genetic Algorithm" if self.type0.isChecked() else "Firefly Algorithm" if self.type1.isChecked() else "Differential Evolution"
        dlg = Algorithm_options_show(type_num, self.Settings)
        dlg.show()
        if dlg.exec_():
            tablePL = lambda row: dlg.PLTable.cellWidget(row, 1).value()
            self.Settings = {'maxGen':dlg.maxGen.value(), 'report':dlg.report.value(),
                'IMax':tablePL(0), 'IMin':tablePL(1),
                'LMax':tablePL(2), 'LMin':tablePL(3),
                'FMax':tablePL(4), 'FMin':tablePL(5),
                'AMax':tablePL(6), 'AMin':tablePL(7)}
            tableAP = lambda row: dlg.APTable.cellWidget(row, 1).value()
            popSize = dlg.popSize.value()
            if type_num=="Genetic Algorithm":
                self.Settings['algorithmPrams'] = {'nPop':popSize, 'pCross':tableAP(0), 'pMute':tableAP(1), 'pWin':tableAP(2), 'bDelta':tableAP(3)}
            elif type_num=="Firefly Algorithm":
                self.Settings['algorithmPrams'] = {'n':popSize, 'alpha':tableAP(0), 'betaMin':tableAP(1), 'gamma':tableAP(2), 'beta0':tableAP(3)}
            elif type_num=="Differential Evolution":
                self.Settings['algorithmPrams'] = {'NP':popSize, 'strategy':tableAP(0), 'F':tableAP(1), 'CR':tableAP(2)}
    
    @pyqtSlot(float)
    def updateRange(self, p0=None):
        self.fixPointRange.emit((self.Ax.value(), self.Ay.value()), self.Ar.value(), (self.Dx.value(), self.Dy.value()), self.Dr.value())
    
    def clear(self):
        self.Point_list.clear()
        self.Result_list.clear()
        self.Settings = self.defaultSettings
        self.X_coordinate.setValue(0)
        self.Y_coordinate.setValue(0)
        self.Ax.setValue(0)
        self.Ay.setValue(0)
        self.Ar.setValue(100)
        self.Dx.setValue(0)
        self.Dy.setValue(0)
        self.Dr.setValue(100)
